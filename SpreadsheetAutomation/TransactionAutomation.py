import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    ws = wb["Sheet1"]

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = ws.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(ws, min_row=2, max_row=ws.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    ws.add_chart(chart, "E2")

    wb.save(filename)

process_workbook("transactions.xlsx")
